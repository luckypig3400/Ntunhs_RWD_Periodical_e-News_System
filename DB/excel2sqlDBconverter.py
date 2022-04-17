from audioop import mul
from types import coroutine
import openpyxl
import mysql.connector
from mysql.connector import errorcode

# 資料庫參數設定
config = {
    "host": "localhost",
    "user": "root",
    "password": "",
    "db": "periodicaldirectConverter"
}

try:
    conn = mysql.connector.connect(
        host=config["host"],
        user=config["user"],
        password=config["password"]
    )

    if input("是否要刪除舊資料庫？(y/n):") == "y":
        cursor = conn.cursor()
        cursor.execute("DROP DATABASE IF EXISTS {}".format(config["db"]))
        conn.commit()
        print("資料庫已刪除")

    # https://www.w3schools.com/python/python_mysql_create_db.asp
    # Auto-create database
    cursor = conn.cursor()
    cursor.execute(
        "CREATE DATABASE IF NOT EXISTS {} DEFAULT CHARACTER SET 'utf8'".format(config["db"]))
    print("Database '{}' established".format(config["db"]))
    cursor.close()
    conn.close()
    conn = mysql.connector.connect(**config)

    print("Connection established")

except mysql.connector.Error as err:
    if err.errno == errorcode.ER_ACCESS_DENIED_ERROR:
        print("Something is wrong with the user name or password")
    elif err.errno == errorcode.ER_BAD_DB_ERROR:
        print("Database does not exist")
    else:
        print(err)
else:
    # 詢問是否建立User與Category資料表
    if input("是否要建立User與Category資料表？(y/n):") == "y":
        # https://stackoverflow.com/questions/59848116/how-to-execute-a-sql-file-using-mysql-connector-and-save-it-in-a-database-in-py
        with open('./empty_DB_with_user_and_categoryTables.sql', 'r', encoding="utf-8") as sqlFile:
            for line in sqlFile:
                # https://stackoverflow.com/questions/3277503/how-to-read-a-file-line-by-line-into-a-list
                cursor = conn.cursor()
                cursor.execute(line)
                cursor.close()
        conn.commit()
        print("帶有預設資料的User與Category資料表已建立")
    # 開啟工作簿
    workbook = openpyxl.load_workbook('newspaper.xlsx')
    # 獲取表單
    sheet = workbook['Content']

    # 直接轉換為新系統用的資料表格式
    if(input("是否直接轉換為periodical資料表？(y/n):") == "y"):
        print("開始轉換為periodical資料表...")
        createTableSQLcommand = "CREATE TABLE IF NOT EXISTS `periodical` (`id` int(11) NOT NULL,`periodNumber` varchar(50) DEFAULT NULL,`noYear` varchar(4) DEFAULT NULL,`noMonth` varchar(2) DEFAULT NULL,`categoryID` varchar(50) DEFAULT NULL,`subject` varchar(50) DEFAULT NULL,`writer` varchar(50) DEFAULT NULL,`quillcontent` mediumtext DEFAULT NULL,`cover` text DEFAULT NULL,`clicked` int(11) DEFAULT NULL,`updateTime` datetime DEFAULT current_timestamp()) ENGINE=InnoDB DEFAULT CHARSET=utf8mb4;"
        cursor = conn.cursor()
        cursor.execute(createTableSQLcommand)
        conn.commit()
        cursor.close()

        for currentRow in range(1, sheet.max_row + 1):
            insertSQLcommand = "INSERT INTO `periodical` (`id`,`periodNumber`,`noYear`,`noMonth`,`categoryID`,`subject`,`writer`,`cover`,`clicked`,`updateTime`) VALUES ("
            # quillcontent請使用quillTextConverter.php 配上content Table轉換

            for currentColumn in range(1, sheet.max_column + 1):
                cellData = sheet.cell(
                    row=currentRow, column=currentColumn).value

                if(currentRow == 1):
                    print("第" + str(currentColumn) + "個欄位名稱為：" + str(cellData))
                else:
                    if currentColumn == 1:
                        print("id:", str(cellData))
                        insertSQLcommand += str(cellData) + ","

                    if currentColumn >= 2 and currentColumn <= 7:
                        # periodNumber, noYear, noMonth, categoryID, subject, writer
                        insertSQLcommand += "'" + str(cellData) + "',"

                    if currentColumn == 13:
                        # First photo as Cover image
                        insertSQLcommand += "'" + str(cellData) + "',"
                    
                    if currentColumn == 32:
                        # clicked
                        insertSQLcommand += str(cellData) + ","

                    if currentColumn == 38:
                        # updateTime
                        insertSQLcommand += "'" + str(cellData) + "')"

                        # 已讀取完畢該列所有資料，執行插入資料庫
                        cursor = conn.cursor()
                        cursor.execute(insertSQLcommand)
                        cursor.close()
                        conn.commit()
                    

        print("成功轉換為periodical資料表!")

    else:
        createTableSQLcommand = "CREATE TABLE IF NOT EXISTS `content` (`Serial` int(11) NOT NULL,`NoID` varchar(50) DEFAULT NULL,`NoYear` varchar(4) DEFAULT NULL,`NoMonth` varchar(2) DEFAULT NULL,`NoClass` varchar(50) DEFAULT NULL,`Subject` varchar(50) DEFAULT NULL,`Writer` varchar(50) DEFAULT NULL,`StarID` int(11) DEFAULT NULL,`Email` varchar(50) DEFAULT NULL,`Content1` mediumtext DEFAULT NULL,`Content2` mediumtext DEFAULT NULL,`Content3` mediumtext DEFAULT NULL,`Photo1` varchar(50) DEFAULT NULL,`Photo2` varchar(50) DEFAULT NULL,`Photo3` varchar(50) DEFAULT NULL,`Alt1` varchar(50) DEFAULT NULL,`Alt2` varchar(50) DEFAULT NULL,`Alt3` varchar(50) DEFAULT NULL,`Temp1` varchar(15) DEFAULT NULL,`Temp2` varchar(15) DEFAULT NULL,`Temp3` varchar(15) DEFAULT NULL,`TempColor` varchar(50) DEFAULT NULL,`Link1` mediumtext DEFAULT NULL,`Link2` mediumtext DEFAULT NULL,`Link3` mediumtext DEFAULT NULL,`Linkloc1` varchar(50) DEFAULT NULL,`Linkloc2` varchar(50) DEFAULT NULL,`Linkloc3` varchar(50) DEFAULT NULL,`Linkalt1` varchar(50) DEFAULT NULL,`Linkalt2` varchar(50) DEFAULT NULL,`Linkalt3` varchar(50) DEFAULT NULL,`Counter` int(11) DEFAULT NULL,`Memo1` varchar(50) DEFAULT NULL,`Memo2` varchar(50) DEFAULT NULL,`Memo3` varchar(50) DEFAULT NULL,`Memo4` varchar(50) DEFAULT NULL,`Memo5` varchar(50) DEFAULT NULL,`TheDate` datetime DEFAULT NULL) CHARSET=utf8mb4;"
        for j in range(1, sheet.max_row + 1):
            # loop for each row
            # https://openpyxl.readthedocs.io/en/stable/api/openpyxl.worksheet.worksheet.html#openpyxl.worksheet.worksheet.Worksheet.max_row
            insertSQLcommand = ""

            for i in range(1, sheet.max_column + 1):
                # loop for each column
                cell = sheet.cell(row=j, column=i).value

                # first row, create table
                if j == 1:
                    print(str(i) + ":" + str(cell))
                    # 列印欄位標題
                    if i == sheet.max_column:
                        cursor = conn.cursor()
                        cursor.execute(createTableSQLcommand)
                        print("Table 'content' established")
                        conn.commit()
                        cursor.close()
                # other rows, insert data
                else:
                    if i == 1:
                        print("current process ID:" + str(cell))
                        insertSQLcommand += "INSERT INTO `content` (`Serial`, `NoID`, `NoYear`, `NoMonth`, `NoClass`, `Subject`, `Writer`, `StarID`, `Email`, `Content1`, `Content2`, `Content3`, `Photo1`, `Photo2`, `Photo3`, `Alt1`, `Alt2`, `Alt3`, `Temp1`, `Temp2`, `Temp3`, `TempColor`, `Link1`, `Link2`, `Link3`, `Linkloc1`, `Linkloc2`, `Linkloc3`, `Linkalt1`, `Linkalt2`, `Linkalt3`, `Counter`, `Memo1`, `Memo2`, `Memo3`, `Memo4`, `Memo5`, `TheDate`) VALUES(" + str(cell) + ","

                    # StarID(int)
                    elif i == 8:
                        if str(cell) == "None":
                            insertSQLcommand += "NULL,"
                        else:
                            insertSQLcommand += str(cell) + ","
                    # Counter(int)
                    elif i == 32:
                        insertSQLcommand += str(cell) + ","
                    # 已讀取完該列的所有欄位資料(最後一欄)
                    elif i == sheet.max_column:
                        insertSQLcommand += "'" + str(cell) + "')"
                        cursor = conn.cursor()
                        # print(insertSQLcommand)
                        cursor.execute(insertSQLcommand)
                        print("Inserted", cursor.rowcount, "row(s) of data.")
                        cursor.close()
                        # 每100筆資料提交一次
                        if j % 100 == 0:
                            conn.commit()
                            print("Transaction committed:", j,
                                  "rows have been inserted.")
                    # 其餘中間欄位資料(用''包住代表字串)
                    else:
                        if str(cell) == "None":
                            insertSQLcommand += "NULL,"
                        else:
                            cellData = str(cell)
                            cellData = cellData.replace("'", "\\'")
                            insertSQLcommand += "'" + cellData + "',"

conn.close()
print("Connection closed.")
